"""
Infraestructura de importacion de archivos de PL_SGE.

Lee tablas en CSV o XLSX y devuelve las filas normalizadas junto con los
errores localizados por fila y columna, de modo que cada modulo solo tenga que
declarar como validar y guardar una fila.

Incluye la validacion del archivo cargado (extension, tipo y tamano), que es
comun a cualquier carga de la plataforma.
"""
from __future__ import annotations

import csv
import io
import unicodedata
from dataclasses import dataclass, field

from django.core.exceptions import ValidationError

# --- Limites y tipos permitidos ---------------------------------------------
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000

IMPORT_EXTENSIONS = {".csv", ".xlsx"}
IMPORT_CONTENT_TYPES = {
    "text/csv",
    "application/csv",
    "text/plain",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
}

DOCUMENT_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt",
    ".png", ".jpg", ".jpeg", ".webp", ".zip",
}
MAX_DOCUMENT_BYTES = 10 * 1024 * 1024


def extension_of(name: str) -> str:
    name = (name or "").strip().lower()
    return name[name.rfind("."):] if "." in name else ""


def validate_upload(uploaded, *, allowed_extensions, max_bytes, allowed_content_types=None):
    """
    Valida extension, tipo declarado y tamano de un archivo cargado.

    Eleva ValidationError con el mensaje correspondiente; no confia solo en el
    `content_type` que envia el navegador: la extension tambien se exige.
    """
    if uploaded is None:
        raise ValidationError({"file": "Debe adjuntar un archivo."})

    extension = extension_of(getattr(uploaded, "name", ""))
    if extension not in allowed_extensions:
        permitidas = ", ".join(sorted(allowed_extensions))
        raise ValidationError({"file": f"Extension no permitida ({extension or 'sin extension'}). Use: {permitidas}."})

    size = getattr(uploaded, "size", 0) or 0
    if size <= 0:
        raise ValidationError({"file": "El archivo esta vacio."})
    if size > max_bytes:
        raise ValidationError(
            {"file": f"El archivo supera el limite de {max_bytes // (1024 * 1024)} MB."}
        )

    if allowed_content_types:
        content_type = (getattr(uploaded, "content_type", "") or "").lower()
        if content_type and content_type not in allowed_content_types:
            raise ValidationError({"file": f"Tipo de archivo no permitido: {content_type}."})

    return uploaded


def validate_document_upload(uploaded):
    """Validacion estandar de un soporte documental de la plataforma."""
    return validate_upload(
        uploaded, allowed_extensions=DOCUMENT_EXTENSIONS, max_bytes=MAX_DOCUMENT_BYTES
    )


# --- Lectura de la tabla -----------------------------------------------------
def normalize_header(value) -> str:
    """`Numero de Documento` -> `numero_de_documento` (sin tildes)."""
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return "_".join(part for part in text.replace("-", " ").replace(".", " ").split())


def read_table(uploaded):
    """
    Devuelve (headers, rows). Cada fila es un dict con las cabeceras
    normalizadas y el numero de fila del archivo en `_row`.
    """
    validate_upload(
        uploaded,
        allowed_extensions=IMPORT_EXTENSIONS,
        max_bytes=MAX_IMPORT_BYTES,
        allowed_content_types=IMPORT_CONTENT_TYPES,
    )
    extension = extension_of(uploaded.name)
    raw = uploaded.read()

    if extension == ".xlsx":
        headers, records = _read_xlsx(raw)
    else:
        headers, records = _read_csv(raw)

    if not headers:
        raise ValidationError({"file": "El archivo no tiene fila de encabezados."})
    if len(records) > MAX_IMPORT_ROWS:
        raise ValidationError(
            {"file": f"El archivo tiene {len(records)} filas y el maximo es {MAX_IMPORT_ROWS}."}
        )
    return headers, records


def _read_csv(raw: bytes):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover
        raise ValidationError({"file": "No fue posible leer el archivo: codificacion no reconocida."})

    sample = text[:4096]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)

    headers = []
    records = []
    for index, row in enumerate(reader, start=1):
        if not any(str(cell).strip() for cell in row):
            continue
        if not headers:
            headers = [normalize_header(cell) for cell in row]
            continue
        records.append(_record(headers, row, index))
    return headers, records


def _read_xlsx(raw: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise ValidationError({"file": "El servidor no puede leer archivos XLSX. Cargue el archivo en CSV."})

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active

    headers = []
    records = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row is None or not any(str(cell).strip() for cell in row if cell is not None):
            continue
        if not headers:
            headers = [normalize_header(cell) for cell in row]
            continue
        records.append(_record(headers, list(row), index))
    workbook.close()
    return headers, records


def _record(headers, row, row_number):
    record = {"_row": row_number}
    for position, header in enumerate(headers):
        if not header:
            continue
        value = row[position] if position < len(row) else ""
        record[header] = "" if value is None else str(value).strip()
    return record


# --- Resultado ---------------------------------------------------------------
@dataclass
class RowError:
    row: int
    column: str
    message: str

    def as_dict(self):
        return {"fila": self.row, "columna": self.column, "mensaje": self.message}


@dataclass
class ImportResult:
    """Resumen de una importacion: creados, actualizados, omitidos y errores."""

    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list = field(default_factory=list)
    headers: list = field(default_factory=list)
    total_rows: int = 0
    dry_run: bool = False

    def add_error(self, row, column, message):
        self.errors.append(RowError(row=row, column=column, message=message))

    @property
    def has_errors(self):
        return bool(self.errors)

    def as_dict(self):
        return {
            "success": not self.has_errors,
            "dry_run": self.dry_run,
            "total_rows": self.total_rows,
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": [error.as_dict() for error in self.errors[:200]],
            "error_count": len(self.errors),
            "headers": self.headers,
        }


def require_columns(headers, required, result):
    """Reporta como error de la fila 1 las columnas obligatorias ausentes."""
    missing = [column for column in required if column not in headers]
    for column in missing:
        result.add_error(1, column, "Falta la columna obligatoria en el encabezado.")
    return not missing
