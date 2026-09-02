"""
Infraestructura comun de la API REST de PL_SGE.

Todos los ViewSets de los modulos heredan de BaseModelViewSet, con lo cual
obtienen automaticamente:
  * Autorizacion por modulo y accion (config.permissions)
  * Trazabilidad de created_by / updated_by
  * Registro en la bitacora de auditoria
  * Borrado logico cuando el modelo lo soporta
  * Exportacion a CSV / Excel  ->  GET /api/<recurso>/export/?format=xlsx
  * Accion de aprobacion       ->  POST /api/<recurso>/<id>/approve/
  * Endpoint de opciones       ->  GET /api/<recurso>/options/
"""
from __future__ import annotations

import csv
import datetime as dt
import io

from django.db import models, transaction
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.serializers import ModelSerializer

from config.permissions import (
    ACTION_APPROVE,
    ACTION_EXPORT,
    HasModulePermission,
    get_permission_map,
)


def scope_to_institution(queryset, request):
    """
    Acota la consulta a la institucion en la que se esta trabajando.

    La ruta hacia la institucion la descubre `core.institutions.scoping` para
    cada modelo, de modo que el aislamiento vive en un solo punto y no hay que
    repetirlo en cada ViewSet.
    """
    from core.institutions.scoping import scope_queryset

    institution = getattr(request, "institution", None)
    user = getattr(request, "user", None)
    return scope_queryset(queryset, institution, user=user)


class AuditModelSerializer(ModelSerializer):
    """Serializer base con metadatos de auditoria en solo lectura."""

    class Meta:
        model = None
        fields = "__all__"
        read_only_fields = ("created_at", "updated_at", "created_by", "updated_by", "uuid")

    def get_request_user(self):
        request = self.context.get("request")
        return getattr(request, "user", None) if request else None


class ExportMixin:
    """
    Exportacion a CSV / Excel compartida por los ViewSets de lectura y de
    escritura, para que un recurso de solo lectura (como la bitacora) exporte
    con el mismo formato y la misma trazabilidad que los demas.
    """

    export_fields: tuple = ()
    export_filename: str = "reporte"

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request, *args, **kwargs):
        self.required_action = ACTION_EXPORT
        queryset = self.filter_queryset(self.get_queryset())
        export_format = request.query_params.get("format", "csv").lower()
        fields = self.export_fields or [
            f.name for f in queryset.model._meta.fields if f.name not in ("id", "uuid")
        ]
        rows = []
        for obj in queryset[:20000]:
            row = []
            for field_name in fields:
                value = obj
                for part in field_name.split("__"):
                    value = getattr(value, part, "") if value is not None else ""
                if callable(value):
                    value = value()
                row.append("" if value is None else str(value))
            rows.append(row)

        self.log_export(queryset.model, len(rows))

        headers = [self._label_for(queryset.model, f) for f in fields]
        if export_format in ("xlsx", "excel"):
            return self._xlsx_response(headers, rows)
        return self._csv_response(headers, rows)

    def log_export(self, model, total):
        """Deja constancia de la exportacion en la bitacora."""
        from core.audit.services import register_audit

        request = self.request
        request._audit_registered = True
        register_audit(
            user=request.user if request.user.is_authenticated else None,
            action="EXPORT",
            module=getattr(self, "module_code", None) or model._meta.app_label,
            request=request,
            description=f"Exportacion de {model._meta.verbose_name_plural}: {total} registros",
        )

    def _label_for(self, model, field_name):
        """Construye la etiqueta de la columna recorriendo las relaciones."""
        parts = field_name.split("__")
        labels = []
        current = model
        try:
            for part in parts:
                field = current._meta.get_field(part)
                labels.append(str(getattr(field, "verbose_name", part)))
                if getattr(field, "related_model", None) is not None:
                    current = field.related_model
            return " / ".join(label.capitalize() for label in labels)
        except Exception:
            return field_name.replace("__", " / ").replace("_", " ").capitalize()

    def _csv_response(self, headers, rows):
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)
        response = HttpResponse(buffer.getvalue().encode("utf-8-sig"), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{self.export_filename}.csv"'
        return response

    def _xlsx_response(self, headers, rows):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return self._csv_response(headers, rows)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.export_filename[:30] or "Reporte"
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            sheet.append(row)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 48)
        sheet.freeze_panes = "A2"

        stream = io.BytesIO()
        workbook.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{self.export_filename}.xlsx"'
        return response


class BaseModelViewSet(ExportMixin, viewsets.ModelViewSet):
    module_code: str | None = None
    permission_classes = [IsAuthenticated, HasModulePermission]
    export_fields: tuple = ()
    export_filename: str = "reporte"
    search_fields: tuple = ()
    ordering_fields = "__all__"
    approve_field: str | None = None
    approve_value = True

    # ------------------------------------------------------------------
    # Consulta
    # ------------------------------------------------------------------
    def get_queryset(self):
        queryset = super().get_queryset()
        model = queryset.model
        if hasattr(model, "deleted_at") and not self.request.query_params.get("include_deleted"):
            queryset = queryset.filter(deleted_at__isnull=True)
        only_active = self.request.query_params.get("only_active")
        if only_active in ("1", "true", "True") and hasattr(model, "is_active"):
            queryset = queryset.filter(is_active=True)
        return scope_to_institution(queryset, self.request)

    # ------------------------------------------------------------------
    # Escritura + auditoria
    # ------------------------------------------------------------------
    def _stamp(self, serializer, creating: bool):
        model = serializer.Meta.model
        payload = {}
        user = self.request.user if self.request.user.is_authenticated else None
        if creating and hasattr(model, "created_by"):
            payload["created_by"] = user
        if hasattr(model, "updated_by"):
            payload["updated_by"] = user
        return payload

    @transaction.atomic
    def perform_create(self, serializer):
        instance = serializer.save(**self._stamp(serializer, creating=True))
        self.log_action("CREATE", instance)
        return instance

    @transaction.atomic
    def perform_update(self, serializer):
        before = self._snapshot(serializer.instance)
        instance = serializer.save(**self._stamp(serializer, creating=False))
        self.log_action("UPDATE", instance, before=before)
        return instance

    @transaction.atomic
    def perform_destroy(self, instance):
        self.log_action("DELETE", instance, before=self._snapshot(instance))
        if hasattr(instance, "soft_delete"):
            user = self.request.user if self.request.user.is_authenticated else None
            instance.soft_delete(user=user)
        else:
            instance.delete()

    # ------------------------------------------------------------------
    # Auditoria
    # ------------------------------------------------------------------
    @staticmethod
    def _snapshot(instance):
        data = {}
        for field in instance._meta.fields:
            value = getattr(instance, field.attname, None)
            if isinstance(value, (dt.date, dt.datetime)):
                value = value.isoformat()
            elif value is not None and not isinstance(value, (str, int, float, bool)):
                value = str(value)
            data[field.name] = value
        return data

    def log_action(self, action_name, instance, before=None):
        from core.audit.services import register_audit

        # Evita que AuditMiddleware duplique el registro con menos detalle.
        self.request._audit_registered = True

        register_audit(
            user=self.request.user if self.request.user.is_authenticated else None,
            action=action_name,
            module=self.module_code or instance._meta.app_label,
            instance=instance,
            request=self.request,
            changes={"before": before} if before else None,
        )

    # ------------------------------------------------------------------
    # Acciones adicionales
    # ------------------------------------------------------------------
    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, *args, **kwargs):
        self.required_action = ACTION_APPROVE
        instance = self.get_object()
        field = self.approve_field
        if not field or not hasattr(instance, field):
            return Response(
                {"success": False, "detail": "Este recurso no admite aprobacion."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        setattr(instance, field, self.approve_value)
        if hasattr(instance, "approved_by"):
            instance.approved_by = request.user
        if hasattr(instance, "approved_at"):
            instance.approved_at = dt.datetime.now(dt.timezone.utc)
        instance.save()
        self.log_action("APPROVE", instance)
        return Response({"success": True, "detail": "Registro aprobado correctamente."})

    @action(detail=True, methods=["post"], url_path="toggle-active")
    def toggle_active(self, request, *args, **kwargs):
        instance = self.get_object()
        if not hasattr(instance, "is_active"):
            return Response(
                {"success": False, "detail": "Recurso sin estado activo/inactivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        instance.is_active = not instance.is_active
        instance.save(update_fields=["is_active"])
        self.log_action("TOGGLE", instance)
        return Response({"success": True, "is_active": instance.is_active})

    @action(detail=False, methods=["get"], url_path="options")
    def options_list(self, request, *args, **kwargs):
        """Devuelve pares id/label para poblar selects del frontend."""
        queryset = self.filter_queryset(self.get_queryset())
        label_field = request.query_params.get("label")
        data = []
        for obj in queryset[:2000]:
            if label_field and hasattr(obj, label_field):
                label = getattr(obj, label_field)
            else:
                label = str(obj)
            data.append({"id": obj.pk, "label": label})
        return Response({"results": data})

    @action(detail=False, methods=["get"], url_path="permissions")
    def my_permissions(self, request, *args, **kwargs):
        perms = get_permission_map(request.user).get(self.module_code, {})
        return Response({"module": self.module_code, "permissions": perms})

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        payload = {"total": queryset.count()}
        model = queryset.model
        if hasattr(model, "is_active"):
            payload["active"] = queryset.filter(is_active=True).count()
            payload["inactive"] = queryset.filter(is_active=False).count()
        return Response(payload)


class ReadOnlyBaseViewSet(ExportMixin, viewsets.ReadOnlyModelViewSet):
    module_code: str | None = None
    permission_classes = [IsAuthenticated, HasModulePermission]

    def get_queryset(self):
        return scope_to_institution(super().get_queryset(), self.request)


def annotate_count(queryset, field, alias=None):
    return queryset.annotate(**{alias or f"{field}_count": models.Count(field)})
